from selenium import webdriver
import time
import re
import openpyxl

driver=webdriver.Firefox(executable_path=r'D:\Python\Python37\Project\OpenSea\firefoxDriver\geckodriver.exe')
driver.get('https://opensea.io/assets?search[query]=Catalog-lu-store')
driver.implicitly_wait(5)    #获取驱动，进入要操作的界面

response=[]
res=[]
for i in range(2,99):
    js = "var q=document.documentElement.scrollTop={}".format(i*100)
    driver.execute_script(js)
    time.sleep(0.01)
    print(driver.page_source)
    response.append(driver.page_source)
regular=r'<a class="sc-1pie21o-0 elyzfO Asset--anchor" href="(.+?)">'
for j in response:
    res.append(re.findall(regular,j))
#试了很多的方法，想用一个变量K循环嵌套列表j[0：len(j)+1]，再通过正则表达是findall的方法换行输入，一个一个提取url地址
#可是输出的结果确实一行行空的列表。
print('-------------------------------------------------------------------------------------')



workbook=openpyxl.Workbook()
sheet1=workbook.create_sheet('sheet1')
sheet1.cell(row=1,colum=1).value='NO'
sheet1.cell(row=1,colum=2).value='URL'
sheet1.cell(row=1,colum=3).value='Status'
workbook.save('data.xlsx')
# 只知道给列表开头

'''
不好意思，写入和导出数据还不会写。
后面我的我就写一下思路。
在excel表中读取一条url数据，调用driver，进入到这个NFT页面，
使用xpath定位到刷新的元素.click()的方法
使用if分支 driver.page_source.__contains__('click')
在Status列的url行输入click。
elif  driver.page_source.__contains__('queued')
在Status列的url行输入queued
else
在Status列的url行输入Error
这很有趣，希望以后能流畅的使用做出来。
如果可以的话我希望能和你学习到更多。
'''